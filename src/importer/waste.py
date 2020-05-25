from logging import getLogger
from openpyxl import load_workbook
import csv
from common.utils import get_tqdm
from importer.base import Importer
logger = getLogger(__name__)

interval_types = {'unknown' : -10, 
                  'on_request' : -11}

def DayStringToNumber(day):
    day_codes = {'po' : 1,
            'út' : 2,
            'st' : 4,
            'čt' : 8,
            'pá' : 16,
            'so' : 32,
            'ne' : 64
           }

    tmp = day.split()
    if(len(tmp) == 1):
        d = tmp[0].strip()[:2].lower()
        return day_codes.get(d, None), day_codes.get(d, None)
    else:
        parity = tmp[0].strip()[:2].lower()
        d = tmp[1].strip()[:2].lower()
        if parity == 'su':
            return day_codes.get(d, 0), 0
        elif parity == 'li':
            return 0, day_codes.get(d, 0)
        else:
            return None, None

class Cheb(Importer):
    def __init__(self):
        super().__init__()
        self.source = "Cheb"

    def Import(self, file):
        super().Import()

        self.workbook = load_workbook(file, read_only = True)

        #load values from xls
        for sheet in self.workbook:
            try:          
                data = self._ParseSheet(sheet)
            except KeyboardInterrupt:
                self.RemoveDBSession()
                logger.error("KeyboardInterrupt")
                raise KeyboardInterrupt
            except e:
                self.RemoveDBSession()
                logger.error("Exception reading source data: %s" % str(e))
                return
            #save to database
            self.SaveAllRecordsToDatabase(data)
            data.clear()

        #get location
        self.LoadOSMLocation()

    def _ParseSheet(self, sheet):
        from models.waste import Cheb
        normalize = {'Typ kontejneru' : 'capacity',
                'Interval' : 'interval',
                'Kód odpadu' : 'waste_code',
                'Název odpadu' : 'waste_name',
                '#TECH_GRP#' : 'tech_grp',
                'Počet ks' : 'quantity',
                'MJ' : 'quantity_unit',
                'Zahájení' : 'start',
                'Ukončení' : 'end',
                'Stav' : 'state',
                'Fakturovat' : 'invoicing',
                'V trase' : 'in_route',
                'Město' : 'city',
                'Ulice' : 'street',
                'č. p.' : 'house_number',
                'Jméno a příjmení' : 'name',
                'Poznámka pro dispečera' : 'note',
                'Poznámka do faktury' : 'invoice_note',
                'Poznámka' : 'dispatcher_note',
                'Den svozu' : 'days_orig',
                'POZNÁMKY PRO OPRAVY' : 'fix_note',
                'Řádek' : 'row',
                'Pořadí' : 'order'}
        data = []
        rows = sheet.rows
        first_row = [cell.value for cell in next(rows)]
        for row in get_tqdm(rows, self.SetState, desc="loading " + sheet.title, total=sheet.max_row):
            record = { 'waste_type': sheet.title }
            for key, cell in zip(first_row, row):
                if cell.data_type == 's':
                    record[normalize.get(key, key)] = cell.value.strip()
                else:
                    record[normalize.get(key, key)] = cell.value

            #Interval
            if record.get('interval'):
                tmp = record.get('interval').split('x')
                if record.get('interval') == 'komplex':
                    record['interval'] = interval_types['unknown']
                elif record.get('interval') == 'na výzvu':
                    record ['interval'] = interval_types['on_request']
                elif len(tmp) == 2:
                    record['interval'] = int(tmp[0]) * (7/int(tmp[1]))
                else:
                    record['interval'] = interval_types['unknown']

            #Days
            record['days_even'] = record['days_odd'] = None
            if record.get('days_orig'):
                even = odd = 0
                tmp = record.get('days_orig').split(',')
                for day in tmp:
                    e, o = DayStringToNumber(day)
                    if e: even += e 
                    if o: odd += o
                record['days_even'] = even
                record['days_odd'] = odd
                            

            data.append(Cheb.as_unique(self.db_session, db_session=self.db_session, data=record))
        return data

class Jihlava(Importer):
    def __init__(self):
        super().__init__()
        self.source = "Jihlava"

    def Import(self, file):
        super().Import()
        if hasattr(file, 'read'):
            filename = file.name
            file.close()
        else:
            filename = file

        data = None
        try:
            if "komunal" in filename:
                data = self._ParseMunicipal(filename)
            elif "separ_hnizda" in filename:
                data = self._ParseSepar(filename)
        except KeyboardInterrupt:
            self.RemoveDBSession()
            logger.error("KeyboardInterrupt")
            raise KeyboardInterrupt
        except Exception as e:
            self.RemoveDBSession()
            logger.error("Exception reading source data: %s" % str(e))
            return

        self.SaveAllRecordsToDatabase(data)
        #get location
        self.LoadOSMLocation()

    def _ParseMunicipal(self, filename):
        from dbfread import DBF
        from pyproj import Proj, transform
        from models.waste import Jihlava
        table = DBF(filename, encoding="cp1250")
        sjtsk = Proj("+init=epsg:5514")
        wgs84= Proj("+init=EPSG:4326")

        data = []
        for row in get_tqdm(table, self.SetState, desc="loading komunal", total=None):
            if not row.get('ADRESA'):
                continue
            record = {}
            record['object_id'] = int(row.get('OBJECTID', -1))
            adresa = row.get('ADRESA').rsplit(' ', 1)
            if len(adresa) > 1 and adresa[1][0].isdigit():
                    record['street'], record['house_number'] = adresa
            else:
                record['street'] = row.get('ADRESA')
                record['house_number'] = ''
            record['city'] = 'Jihlava'
            record['country'] = 'Czech republic'
            record['longitude'], record['latitude'] = transform(sjtsk, wgs84, row.get('X'), row.get('Y'))
            record['population'] = row.get('OBYVATEL')
            record['name'] = row.get('NAZEV')
            record['waste_type'] = 'TKO'
            record['waste_code'] = '200301'
            record['waste_name'] = 'Směsný komunální odpad'

            record['quantity'] = row.get('POCET')
            record['quantity_unit'] = 'ks'
            record['capacity'] = row.get('NADOBA')

            record['start'] = row.get('OD_DATUM')
            record['end'] = row.get('DO_DATUM')
            record['interval'] = row.get('CETNO_SVOZ')
            record['days_orig'] = None
            record['days_odd'] = record['days_even'] = row.get('SVOZ_PO') * 1 + \
                                                       row.get('SVOZ_UT') * 2 + \
                                                       row.get('SVOZ_ST') * 4 + \
                                                       row.get('SVOZ_CT') * 8 + \
                                                       row.get('SVOZ_PA') * 16 

            record['note'] = row.get('POZNAMKA')

            record['ownership'] = row.get('TYP_VLASTN')
            record['optimum'] = int(row.get('OPTIMUM', -1))
            record['coefficient'] = int(row.get('KOEFICIENT', -1))
            record['ratio'] = int(row.get('POMER', -1))
            #TODO: POCET1, NADOBA1, CETN_SVOZ1
            data.append(Jihlava.as_unique(self.db_session, db_session=self.db_session, data=record))

        return data

    def _ParseSepar(self, filename):
        raise Exception("Separ is not implemented")

        from dbfread import DBF
        table = DBF(filename, encoding="cp1250")

        data = []
        for row in get_tqdm(table, self.SetState, desc="loading separ_hnizda", total=None):
            record = {}
            data.append(record)

class Stavanger(Importer):
    def __init__(self):
        super().__init__()
        self.source = "Stavanger"

    def Import(self, file):
        super().Import()

        data = None
        try:
            data = self._ParseCSV(file)
        except KeyboardInterrupt:
            self.RemoveDBSession()
            logger.error("KeyboardInterrupt")
            raise KeyboardInterrupt
        except Exception as e:
            self.RemoveDBSession()
            logger.error("Exception reading source data: %s" % str(e))
            return
        self.SaveAllRecordsToDatabase(data)
        self.LoadOSMLocation()

    def _ParseCSV(self, file):
        from models.waste import Stavanger
        data = []

        waste_codes = {'Bio' : 200201,
                        'Bio-næring' : 200201,
                        'Glass/metall' : 200102,
                        'Hytter Papir' : 200101,
                        'Hytter Rest' : 200301,
                        'Papir' : 200101,
                        'Papir-næring' : 200101,
                        'Plast' : 200139,
                        'Plast-næring' : 200139,
                        'Restavfall' : 200301,
                        'Rest-næring' : 200301,
                        'Rest-vask' : 200301,
                        'Tekstil' : 200111}

        with open(file, encoding='utf-8') as csvfile:
            numline = len(csvfile.readlines())
            csvfile.seek(0)
            reader = csv.DictReader(csvfile, delimiter=',', quotechar='\'', quoting=csv.QUOTE_ALL)
            for row in get_tqdm(reader, self.SetState, desc="loading Stavanger csv", total=numline):
                if not row['LATITUDE'] or not row['LONGITUDE']:
                    continue
                record = {}
                
                record['object_id'] = int(row['CONTAINERNUMBER'])

                record['street'] = row['ADDRESS']
                record['house_number'] = ''

                record['city'] = 'Stavanger'
                record['country'] = 'Norway'
                record['latitude'] = float(row['LATITUDE'])
                record['longitude'] = float(row['LONGITUDE'])
                record['waste_name'] = row['FRACTION']

                record['name'] = None
                record['waste_type'] = None
                record['waste_code'] = waste_codes.get(row['FRACTION'], 0)
                record['quantity'] = None
                record['quantity_unit'] = None
                record['capacity'] = None
                record['days_orig'] = record['days_even'] = record['days_odd'] = None
                record['start'] = None
                record['end'] = None
                record['interval'] = None
                record['note'] = None

                record['counter'] = int(row['COUNTER'])
                record['fillheight'] = int(row['FILLHEIGHT'])
                record['date'] = row['DATE']

                data.append(Stavanger.as_unique(self.db_session, db_session=self.db_session, data=record))
        return data

class Plzen(Importer):
    def __init__(self):
        super().__init__()
        self.source = "Plzen"

    def Import(self, file):
        super().Import()

        self.workbook = load_workbook(file, read_only = True)

        #load values from xls
        for sheet in self.workbook:
            try:
                data = self._ParseSheet(sheet)
            except KeyboardInterrupt:
                self.RemoveDBSession()
                logger.error("KeyboardInterrupt")
                raise KeyboardInterrupt
            except e:
                self.RemoveDBSession()
                logger.error("Exception reading source data: %s" % str(e))
                return
            #save to database
            self.SaveAllRecordsToDatabase(data)
            data.clear()

        #get location
        self.LoadOSMLocation(city_filter=self.source)

    def _ParseSheet(self, sheet):
        from models.waste import Plzen
        normalize = {'ID_CONTAINER' : 'object_id',
                     'TrashType' : 'waste_name',
                     'Volume' : 'capacity',
                     'Latitude' : 'latitude',
                     'Longitude' : 'longitude',
                     'Interval' : 'interval',
                     'CollectionPlaceName' : 'address',
                     'Variant' : 'variant',
                     'Count': 'quantity',
                     'ID_COLLECTION_PLACE': 'collection_place',
                     'ID_DISTRICT': 'district'
                }

        waste_codes = {'MixedWaste': 200301,
                       'Paper' : 200101,
                       'BioWaste': 200201,
                       'Glass': 200102,
                       'Plastics': 200139
                       }

        data = []
        rows = sheet.rows
        first_row = [cell.value for cell in next(rows)]
        for row in get_tqdm(rows, self.SetState, desc="loading " + sheet.title, total=sheet.max_row):
            record = { 'waste_type': sheet.title }
            for key, cell in zip(first_row, row):
                if cell.data_type == 's':
                    record[normalize.get(key, key)] = cell.value.strip()
                else:
                    record[normalize.get(key, key)] = cell.value

            # upraveni adresy
            adresa = record.get('address').strip('/ ').rsplit(' ', 1)
            if len(adresa) > 1 and adresa[1][0].isdigit():
                    record['street'], record['house_number'] = adresa
            else:
                record['street'] = record.get('address')
                record['house_number'] = ''
            record.pop('address', None)

            record['waste_code'] = waste_codes.get(record.get('waste_name', ''), 0)
            record['city'] = self.source
            record['country'] = 'Czech republic'

            record['name'] = None
            record['waste_type'] = None
            record['quantity'] = record.get('quantity', None)
            record['quantity_unit'] = None
            record['days_orig'] = record['days_even'] = record['days_odd'] = None
            record['start'] = None
            record['end'] = None
            record['note'] = None
            record['variant'] = record.get('variant', None)
            record['collection_place'] = record.get('collection_place', None)
            record['district'] = record.get('district', None)

            data.append(Plzen.as_unique(self.db_session, db_session=self.db_session, data=record))
        return data


class Tabor(Importer):
    def __init__(self):
        super().__init__()
        self.source = "Tabor"

    def Import(self, file):
        super().Import()

        self.workbook = load_workbook(file, read_only = True)

        #load values from xls
        for sheet in self.workbook:
            try:
                data = self._ParseSheet(sheet)
            except KeyboardInterrupt:
                self.RemoveDBSession()
                logger.error("KeyboardInterrupt")
                raise KeyboardInterrupt
            except e:
                self.RemoveDBSession()
                logger.error("Exception reading source data: %s" % str(e))
                return
            #save to database
            self.SaveAllRecordsToDatabase(data)
            data.clear()

        #get location
        self.LoadOSMLocation(city_filter=self.source)

    def _ParseSheet(self, sheet):
        from models.waste import Tabor
        normalize = {'ID_CONTAINER' : 'object_id',
                     'TrashType' : 'waste_name',
                     'Volume' : 'capacity',
                     'Latitude' : 'latitude',
                     'Longitude' : 'longitude',
                     'Interval' : 'interval',
                     'CollectionPlaceName' : 'address',
                     'Variant' : 'variant',
                     'quantity': 'quantity'
                }
        
        waste_codes = {'Paper' : 200101,
                       'MMW': 200301,
                       'Bio': 200201,
                       'Glass': 200102,
                       'Plastic': 200139,}
        data = []
        rows = sheet.rows
        first_row = [cell.value for cell in next(rows)]
        for row in get_tqdm(rows, self.SetState, desc="loading " + sheet.title, total=sheet.max_row):
            record = { 'waste_type': sheet.title }
            for key, cell in zip(first_row, row):
                if cell.data_type == 's':
                    record[normalize.get(key, key)] = cell.value.strip()
                else:
                    record[normalize.get(key, key)] = cell.value

            # upraveni adresy
            adresa = record.get('address').rsplit(' ', 1)
            if len(adresa) > 1 and adresa[1][0].isdigit():
                    record['street'], record['house_number'] = adresa
            else:
                record['street'] = record.get('address')
                record['house_number'] = ''
            record.pop('address', None)
            #record['street'] = None
            #record['house_number'] = ''

            record['waste_code'] = waste_codes.get(record.get('waste_name', ''), 0)
            record['city'] = self.source
            record['country'] = 'Czech republic'

            record['name'] = None
            record['waste_type'] = None
            record['quantity'] = record.get('quantity', None)
            record['quantity_unit'] = None
            record['days_orig'] = record['days_even'] = record['days_odd'] = None
            record['start'] = None
            record['end'] = None
            record['note'] = None
            record['variant'] = record.get('variant', None)

            data.append(Tabor.as_unique(self.db_session, db_session=self.db_session, data=record))
        return data