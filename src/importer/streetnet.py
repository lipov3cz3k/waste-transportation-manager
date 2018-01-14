from importer.base import Importer
from openpyxl import load_workbook
from logging import getLogger
from tqdm import tqdm
LOGGER = getLogger(__name__)
from models.cdv import StreetnetSegments

class StreetNet(Importer):
    def __init__(self):
        super().__init__()
        self.source = "StreetNet"


    def Import(self, file):
        super().Import()

        self.workbook = load_workbook(file, read_only = True)

        #load values from xls
        for sheet in self.workbook:
            try:          
                data = self._ParseSheet(sheet)
            except KeyboardInterrupt:
                self.RemoveDBSession()
                LOGGER.error("KeyboardInterrupt")
                raise KeyboardInterrupt
            except Exception as e:
                self.RemoveDBSession()
                LOGGER.error("Exception reading source data: %s" % str(e))
                return
            #save to database
            self.SaveAllRecordsToDatabase(data)
            data.clear()


    def _ParseSheet(self, sheet):
        normalize = {'ROAD_ID' : 'id',
                     'delka' : 'length',
                     'cas' : 'time',
                     'silnice' : 'way_name',
                     'trida_kom' : 'way_class',
                     'most' : 'bridge',
                     'podjezd' : 'underpass',
                     'tunel' : 'tunnel',
                     'MIN_nosnost' : 'min_load_capacity',
                     'MIN_vyska' : 'min_height',
                     'X_start' : 'start_X',
                     'Y_start' : 'start_Y',
                     'X_end' : 'end_X',
                     'Y_end' : 'end_Y'}
        data = []
        rows = sheet.rows
        first_row = [cell.value for cell in next(rows)]
        for row in tqdm(rows, desc="loading " + sheet.title, total=sheet.max_row):
            record = {}
            for key, cell in zip(first_row, row):
                if cell.data_type == 's':
                    record[normalize.get(key, key)] = cell.value.strip()
                else:
                    record[normalize.get(key, key)] = cell.value
            data.append(StreetnetSegments.as_unique(self.db_session, **record))
        return data
