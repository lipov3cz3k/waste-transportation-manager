from openpyxl import load_workbook
from datetime import datetime, time
from common.config import local_config
from common.utils import LogType, print, _print, get_tqdm
from waste.importer import Importer
from models.tracks import Track
from .routing import RoutingType
from re import split
from multiprocessing.pool import ThreadPool
from sqlalchemy import inspect
from sqlalchemy.orm import Session

def chunks(l, n):
    """Yield successive n-sized chunks from l."""
    for i in range(0, len(l), n):
        yield l[i:i + n]

class TrackImporter(Importer):
    def __init__(self):
        super().__init__()
        self.source = "tracks"

    def Import(self, file):
        super().Import()

        self.workbook = load_workbook(file, read_only = True)
        #load values from xls
        for sheet in self.workbook:
            try:          
                data = self._ParseSheet(sheet)
            except KeyboardInterrupt:
                self.RemoveDBSession()
                print("KeyboardInterrupt", LogType.error)
                raise KeyboardInterrupt
            except Exception as e:
                self.RemoveDBSession()
                print("Exception reading source data: %s" % str(e), LogType.error)
                return
            #save to database
            self.SaveAllRecordsToDatabase(data)
            data.clear()
        #get location
        self.LoadOSMLocation()

    def _ParseSheet(self, sheet):
        normalize = {
                'od' : 'date_from',
                'do' : 'date_to',
                'vozidlo' : 'vehicle',
                # Vodafone
                'start' : 'start',
                'cíl' : 'finish',
                'SPZ' : 'reg_plate',
                'řidič' : 'driver',
                'popis' : 'note',
                'typ' : 'type',
                'tank' : 'tank',
                'vzdálenost metrů' : 'distance',
                'čas' : 'time',
                'Tach. start' : 'tachometer_start',
                'Tach.  cíl' : 'tachometer_finish',
                'stejné' : 'same',
                # Positrex
                'Začátek' : 'start',
                'Konec' : 'finish',
                'Vzdalenost [m]' : 'distance',
                'Doba jízdy [min]' : 'time',
                'Prům. rychlost [km/h]' : 'avg_speed',
                'Max. rychlost [km/h]' : 'max_speed',
                'Poznámka 1' : 'note',
                'Poznámka 2' : 'note2'}

        data = []
        rows = sheet.rows
        first_row = [cell.value for cell in next(rows)]
        for row in get_tqdm(rows, self.SetState, desc="loading " + sheet.title, total=sheet.max_row):
            record = {}
            for key, cell in zip(first_row, row):
                if cell.data_type == 's':
                    record[normalize.get(key, key)] = cell.value.strip()
                else:
                    record[normalize.get(key, key)] = cell.value
            if record.get('note2') == "Positrex":
                record['start_address'] = self._ParsePositrexAddress(record.get('start'))
                record['finish_address'] = self._ParsePositrexAddress(record.get('finish'))
                record['date_from'] = datetime.strptime(record['date_from'], '%d.%m.%Y %H:%M')
                record['date_to'] = datetime.strptime(record['date_to'], '%d.%m.%Y %H:%M')
                record['time'] = datetime.strptime('{:02d}:{:02d}'.format(*divmod(int(round(record['time'])), 60)), '%H:%M').time()
            else:
                record['start_address'] = self._ParseVodafoneAddress(record.get('start'))
                record['finish_address'] = self._ParseVodafoneAddress(record.get('finish'))
            data.append(Track.as_unique(self.db_session, db_session=self.db_session, data=record))
        return data

    def _ParseVodafoneAddress(self, address):
        result = split('^(.*),\s?(\d{3}\s\d{2})\s*(.*),\s*(.*)', address)
        if len(result) == 1:
            adresa = address.split(',')
            if len(adresa) < 2:
                raise
            result.insert(1, adresa[0])
            result.insert(2, '')
            result.insert(3, adresa[-2])
            result.insert(4, adresa[-1])

        adresa = result[1].rsplit(' ', 1)
        if len(adresa) > 1 and adresa[1][0].isdigit():
                street, house_number = adresa
        else:
            street = result[1]
            house_number = ''

        return {'street': street, 'house_number' : house_number, 'postal' : result[2].replace(' ', ''), 'city' : result[3], 'country' : result[4]}

    def _ParsePositrexAddress(self, address):
        parts = [x.strip() for x in address.split(',')]
        result = {}
        result['country'] = parts[0]
        result['city'] = parts[1]
        #result['postal'] = parts[0]
        if len(parts) > 2:
            if parts[-1][0].isdigit():
                result['house_number'] = parts[-1]
                if len(parts) > 3:
                    result['street'] = parts[-2]
            else:
                result['street'] = parts[-1]
        return result

    def GetTracks(self, bbox):
        from database import init_db, db_session
        from models.location import Address, OSMLocation
        from models.path import Path
        from sqlalchemy.orm import aliased
        init_db()
        self.db_session = db_session
        self.TestIsRun()
        start_alias = aliased(Address)
        finish_alias = aliased(Address)
        OSMstart_alias = aliased(OSMLocation)
        OSMfinish_alias = aliased(OSMLocation)
        tracks_obj = db_session.query(Track, OSMstart_alias, OSMfinish_alias).join(start_alias, Track.start_address) \
                                        .join(finish_alias, Track.finish_address) \
                                        .join(OSMstart_alias, start_alias.location) \
                                        .join(OSMfinish_alias, finish_alias.location) \
                                        .outerjoin(Path) \
                                        .filter(Path.track_id == None) \
                                        .filter(OSMstart_alias.latitude > bbox.min_latitude, \
                                                OSMstart_alias.latitude < bbox.max_latitude, \
                                                OSMstart_alias.longitude > bbox.min_longitude, \
                                                OSMstart_alias.longitude < bbox.max_longitude ) \
                                        .all()
        return tracks_obj


class PathFinder:
    def __init__(self, network, state = None, run = {0: False}):
        self.state = state
        self.network = network
        self.run = run
        

    def SetState(self, action = None, percentage = None):
        if action != None:
            self.state['action'] = action
        if percentage != None:
            self.state['percentage'] = int(percentage)

    def TestIsRun(self):
        if not self.run[0]:
            raise Exception("Service not running")


    def GetTracksWithPaths(self, safeToDb = True):
        from .path_finder import TrackImporter
        self.run[0] = True
        importer = TrackImporter()
        tracks_chunks = []

        try:
            if importer:
                importer.run = True
                tracks_chunks = importer.GetTracks(self.network.bbox)
                tracks_chunks = list(chunks(tracks_chunks, 100))
        except KeyboardInterrupt as e:
            importer.run = False
            print("KeyboardInterrupt", LogType.info)
            raise e
        except Exception as e:
            print("Exception reading source data: %s" % str(e), LogType.error)
            return

        thread_results = {}
        thread_pool = ThreadPool(processes=local_config.thread_pool_size_for_mapping)

        for index, tracks_obj in enumerate(tracks_chunks):
            thread_pool.apply_async(self._GetTracksRouting, args=(tracks_obj, safeToDb, index, len(tracks_chunks), thread_results))

        print("Threads started, wait for finish.")

        thread_pool.close()
        thread_pool.join()
        print("Threads finished.")
        return thread_results



    def _searchNearby(self, point):
        from graph.bounding_box import get_bounding_box
        from shapely.geometry import Point as splPoint
        bbox = get_bounding_box(point.y,  point.x, 0.5)

        nodes = (n for n,d in self.network.G.nodes_iter(data=True) if d['lat'] > bbox.min_latitude and \
                                                              d['lat'] < bbox.max_latitude and \
                                                              d['lon'] > bbox.min_longitude  and \
                                                              d['lon'] < bbox.max_longitude)
        nodes = list(nodes)
        if len(nodes) == 0:
            return None
        try:
            dist = lambda node: point.distance(splPoint(self.network.G.node[node]['lon'], self.network.G.node[node]['lat']))
            near_node = min(nodes, key=dist)
            return near_node
        except Exception as e:
            print(str(e), log_type=LogType.error)
            raise e



    def _GetTracksRouting(self, tracks_obj, safeToDb, iteration, iteration_total, thread_results):
        from database import db_session
        local_db_session = db_session()
        from models.path import Path
        from shapely.geometry import Point as splPoint
        result = []
        db_temp = []
        try:
            #for track in get_tqdm(tracks_obj, self.SetState, desc="Connecting tracks to network", total=None):
            for track in get_tqdm(tracks_obj, self.SetState, desc="Connecting tracks to network %d/%d:" % (iteration, iteration_total), position=iteration):
            #for track in tracks_obj:
                self.TestIsRun()
                start_node = self._searchNearby(splPoint(float(track[1].longitude), float(track[1].latitude)))
                finish_node = self._searchNearby(splPoint(float(track[2].longitude), float(track[2].latitude)))

                if not start_node or not finish_node:
                    continue 
                route = self.network.Route(start_node, finish_node)

                if safeToDb:
                    #route['track']  = local_db_session.query(Track).get(track[0])
                    track_session = Session.object_session(track[0])
                    track_session.expunge(track[0])
                    route['track'] = track[0]
                    db_temp.append(Path(db_session=local_db_session, data=route))
                else:
                    route['track'] = track[0]
                    result.append(route)

            if safeToDb:
                local_db_session.add_all(db_temp)
                local_db_session.commit()
        except Exception as e:
            if safeToDb:
                local_db_session.add_all(db_temp)
                local_db_session.commit()
            #local_db_session.close()
            raise e
        _print("Thread %d / %d finished." % (iteration+1, iteration_total))
        thread_results[iteration] = result
        #local_db_session.close()

def Run(importFile=None, sourceCity=None):
    importer = TrackImporter()
    try:
        if importer:
            importer.run = True
            importer.Import(importFile)
    except KeyboardInterrupt:
        importer.run = False
        print("KeyboardInterrupt", LogType.info)
    except Exception as e:
        print("Exception reading track source data: %s" % str(e), LogType.error)