from logging import getLogger
from graph.graph_factory import load

logger = getLogger(__name__)


def trackinfo(graph_file, input_file):
    from openpyxl import load_workbook        
    from models.cdv import StreetnetOSMRelation, StreetnetSegments
    from database import init_db, db_session

    def _get_cdv_info(db_session, edges, methods):
        ids = [e['id'] for e in edges]
        segments = db_session.query(StreetnetSegments).join(StreetnetOSMRelation) \
                                                            .filter(StreetnetOSMRelation.osm_way_id.in_(ids)) \
                                                            .all()
        param, method = methods
        tmp = []
        for segment in segments:
            tmp.append(getattr(segment, param))
        return method(tmp)

    def _process_sheet(sheet, graph):
        from tqdm import tqdm
        rows = sheet.rows
        first_row = [cell.value for cell in next(rows)]

        normalize = {'start' : 'start',
                    'cil' : 'end',
                    'parametr' : 'parameter',
                    'metoda' : 'method',
                    'vysledek' : 'result'}
        allowed_methods = {'min' : min,
                        'max' : max}
        for row_index, row in tqdm(enumerate(rows, start=2), desc="loading " + sheet.title, total=sheet.max_row):
            record = {}
            for key, cell in zip(first_row, row):
                if cell.data_type == 's':
                    record[normalize.get(key, key)] = cell.value.strip()
                else:
                    record[normalize.get(key, key)] = cell.value

            method = allowed_methods.get(record['method'], None)
            if not method:
                raise ValueError("Specified (%s) math method is not supported" % record['method'])
            cdv_params = (record['parameter'], method)
            response = graph.route_by_NUTS5(record['start'], record['end'])
            cdv_data = _get_cdv_info(db_session, response.get('paths').get('features')[0].get('properties').get('edges'), cdv_params) 
            logger.info("Route from %s to %s has parameter %s %s %s", record['start'], record['end'], record['parameter'], record['method'], cdv_data)
            sheet.cell(row=row_index, column=5).value = cdv_data

    init_db()
    graph = load(graph_file)
    try:
        workbook = load_workbook(input_file)
        for sheet in workbook:
            _process_sheet(sheet, graph)
        workbook.save(input_file)
    except Exception as ex:
        logger.exception("Error in loading workbook %s", ex)