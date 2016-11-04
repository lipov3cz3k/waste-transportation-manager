from openpyxl import load_workbook
from common.utils import LogType, print, get_tqdm

interval_types = {'unknown' : -10, 
                  'on_request' : -11}

def DayStringToNumber(day):
    day_codes = {'po' : 1,
            'út' : 2,
            'st' : 4,
            'čt' : 8,
            'pá' : 16,
            'so' : 32,
            'ne' : 64
           }

    tmp = day.split()
    if(len(tmp) == 1):
        d = tmp[0].strip()[:2].lower()
        return day_codes.get(d, None), day_codes.get(d, None)
    else:
        parity = tmp[0].strip()[:2].lower()
        d = tmp[1].strip()[:2].lower()
        if parity == 'su':
            return day_codes.get(d, 0), 0
        elif parity == 'li':
            return 0, day_codes.get(d, 0)
        else:
            return None, None


class Importer:
    def __init__(self):
        self.state = {}
        self.SetState(action="init", percentage=0)
        self.source = None
        self.workbook = None
        self.run = False
        self.db_session = None

    def RemoveDBSession(self):
        if self.db_session:
            self.db_session.remove()

    def SetState(self, action = None, percentage = None):
        if action != None:
            self.state['action'] = action
        if percentage != None:
            self.state['percentage'] = int(percentage)

    def TestIsRun(self):
        if not self.run:
            raise Exception("Service not running")

    def Import(self):
        from database import init_db, db_session
        init_db()
        self.db_session = db_session
        print("start import %s" % self.source)
        
    def SaveAllRecordsToDatabase(self, records):
        print("Saving records to database")#), LogType.trace)
        self.db_session.add_all(records)
        self.db_session.commit()

    def LoadOSMLocation(self):
        print("Start reversing Addresses.")#, LogType.info)
        addresses_without_location = self._GetAddressesWithoutLocation()
        self._GetLocationsForAddresses(addresses_without_location)

    def _GetAddressesWithoutLocation(self):
        from models.location import Address

        addresses_obj = self.db_session.query(Address).filter(Address.location == None).all()
        return addresses_obj

    def _GetLocationsForAddresses(self, addresses_array):
        from time import sleep
        from geopy.geocoders import Nominatim
        from geopy.exc import GeocoderTimedOut
        from models.location import OSMLocation
        print("Getting locations for Addresses <%d> (use Nominatim)" % len(addresses_array))#, LogType.trace)

        geolocator = Nominatim()
        for addr_obj in get_tqdm(addresses_array, self.SetState, desc="getting locations", total=None):
            #print("getting new by nominatim ...", LogType.trace, only_Message=True)
            successful = False
            while not successful:
                location = self.db_session.query(OSMLocation).filter(OSMLocation.road == addr_obj.street,
                                                                     OSMLocation.house_number == addr_obj.house_number,
                                                                     ((OSMLocation.city == addr_obj.city) | (OSMLocation.town == addr_obj.city))).first()
                if location:
                    osm_id=location.osm_id
                    address = []
                    successful = True
                else:
                    try:
                        location = geolocator.geocode({'street':addr_obj.house_number + ' ' + addr_obj.street,'city':addr_obj.city},
                                                      addressdetails=True)
                    except GeocoderTimedOut as e:
                        print("Service timed out, waiting a little bit")
                        sleep(10)
                    except Exception as e:
                        raise e
                    else:
                        successful = True
                    if location == None:
                        continue
                    address = location.raw['address']
                    osm_id = location.raw['osm_id']
                    sleep(1)
                location_obj = OSMLocation.as_unique(self.db_session,
                                                  address=address, 
                                                  osm_id=osm_id, 
                                                  latitude=location.latitude, 
                                                  longitude=location.longitude)
                addr_obj.set_location(location_obj)
                self.db_session.commit()


class Cheb(Importer):
    def __init__(self):
        super().__init__()
        self.source = "Cheb"

    def Import(self, file):
        super().Import()

        self.workbook = load_workbook(file, read_only = True)

        #load values from xls
        for sheet in self.workbook:
            try:          
                data = self._ParseSheet(sheet)
            except KeyboardInterrupt:
                self.RemoveDBSession()
                print("KeyboardInterrupt", LogType.error)
                raise KeyboardInterrupt
            except e:
                self.RemoveDBSession()
                print("Exception reading source data: %s" % str(e), LogType.error)
                return
            #save to database
            self.SaveAllRecordsToDatabase(data)
            data.clear()

        #get location
        self.LoadOSMLocation()

    def _ParseSheet(self, sheet):
        from models.waste import Cheb
        normalize = {'Typ kontejneru' : 'capacity',
                'Interval' : 'interval',
                'Kód odpadu' : 'waste_code',
                'Název odpadu' : 'waste_name',
                '#TECH_GRP#' : 'tech_grp',
                'Počet ks' : 'quantity',
                'MJ' : 'quantity_unit',
                'Zahájení' : 'start',
                'Ukončení' : 'end',
                'Stav' : 'state',
                'Fakturovat' : 'invoicing',
                'V trase' : 'in_route',
                'Město' : 'city',
                'Ulice' : 'street',
                'č. p.' : 'house_number',
                'Jméno a příjmení' : 'name',
                'Poznámka pro dispečera' : 'note',
                'Poznámka do faktury' : 'invoice_note',
                'Poznámka' : 'dispatcher_note',
                'Den svozu' : 'days_orig',
                'POZNÁMKY PRO OPRAVY' : 'fix_note',
                'Řádek' : 'row',
                'Pořadí' : 'order'}
        data = []
        rows = sheet.rows
        first_row = [cell.value for cell in next(rows)]
        for row in get_tqdm(rows, self.SetState, desc="loading " + sheet.title, total=sheet.max_row):
            record = { 'waste_type': sheet.title }
            for key, cell in zip(first_row, row):
                if cell.data_type == 's':
                    record[normalize.get(key, key)] = cell.value.strip()
                else:
                    record[normalize.get(key, key)] = cell.value

            #Interval
            if record.get('interval'):
                tmp = record.get('interval').split('x')
                if record.get('interval') == 'komplex':
                    record['interval'] = interval_types['unknown']
                elif record.get('interval') == 'na výzvu':
                    record ['interval'] = interval_types['on_request']
                elif len(tmp) == 2:
                    record['interval'] = int(tmp[0]) * (7/int(tmp[1]))
                else:
                    record['interval'] = interval_types['unknown']

            #Days
            record['days_even'] = record['days_odd'] = None
            if record.get('days_orig'):
                even = odd = 0
                tmp = record.get('days_orig').split(',')
                for day in tmp:
                    e, o = DayStringToNumber(day)
                    if e: even += e 
                    if o: odd += o
                record['days_even'] = even
                record['days_odd'] = odd
                            

            data.append(Cheb.as_unique(self.db_session, db_session=self.db_session, data=record))
        return data

class Jihlava(Importer):
    def __init__(self):
        super().__init__()
        self.source = "Jihlava"

    def Import(self, file):
        super().Import()
        if hasattr(file, 'read'):
            filename = file.name
            file.close()
        else:
            filename = file

        data = None
        try:
            if "komunal" in filename:
                data = self._ParseMunicipal(filename)
            elif "separ_hnizda" in filename:
                data = self._ParseSepar(filename)
        except KeyboardInterrupt:
            self.RemoveDBSession()
            print("KeyboardInterrupt", LogType.error)
            raise KeyboardInterrupt
        except Exception as e:
            self.RemoveDBSession()
            print("Exception reading source data: %s" % str(e), LogType.error)
            return

        self.SaveAllRecordsToDatabase(data)
        #get location
        self.LoadOSMLocation()

    def _ParseMunicipal(self, filename):
        from dbfread import DBF
        from pyproj import Proj, transform
        from models.waste import Jihlava
        table = DBF(filename, encoding="cp1250")
        sjtsk = Proj("+init=epsg:5514")
        wgs84= Proj("+init=EPSG:4326")

        data = []
        for row in get_tqdm(table, self.SetState, desc="loading komunal", total=None):
            if not row.get('ADRESA'):
                continue
            record = {}
            record['object_id'] = int(row.get('OBJECTID', -1))
            adresa = row.get('ADRESA').rsplit(' ', 1)
            if len(adresa) > 1 and adresa[1][0].isdigit():
                    record['street'], record['house_number'] = adresa
            else:
                record['street'] = row.get('ADRESA')
                record['house_number'] = ''
            record['city'] = 'Jihlava'
            record['latitude'], record['longitude'] = transform(sjtsk, wgs84, row.get('X'), row.get('Y'))
            record['population'] = row.get('OBYVATEL')
            record['name'] = row.get('NAZEV')
            record['waste_type'] = 'TKO'
            record['waste_code'] = '200301'
            record['waste_name'] = 'Směsný komunální odpad'

            record['quantity'] = row.get('POCET')
            record['quantity_unit'] = 'ks'
            record['capacity'] = row.get('NADOBA')

            record['start'] = row.get('OD_DATUM')
            record['end'] = row.get('DO_DATUM')
            record['interval'] = row.get('CETNO_SVOZ')
            record['days_orig'] = None
            record['days_odd'] = record['days_even'] = row.get('SVOZ_PO') * 1 + \
                                                       row.get('SVOZ_UT') * 2 + \
                                                       row.get('SVOZ_ST') * 4 + \
                                                       row.get('SVOZ_CT') * 8 + \
                                                       row.get('SVOZ_PA') * 16 

            record['note'] = row.get('POZNAMKA')

            record['ownership'] = row.get('TYP_VLASTN')
            record['optimum'] = int(row.get('OPTIMUM', -1))
            record['coefficient'] = int(row.get('KOEFICIENT', -1))
            record['ratio'] = int(row.get('POMER', -1))
            #TODO: POCET1, NADOBA1, CETN_SVOZ1
            data.append(Jihlava.as_unique(self.db_session, db_session=self.db_session, data=record))

        return data

    def _ParseSepar(self, filename):
        raise Exception("Separ is not implemented")

        from dbfread import DBF
        table = DBF(filename, encoding="cp1250")

        data = []
        for row in get_tqdm(table, self.SetState, desc="loading separ_hnizda", total=None):
            record = {}
            data.append(record)
