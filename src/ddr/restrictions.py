from logging import getLogger
from openpyxl import load_workbook
from common.utils import get_tqdm
from geojson import Point, Feature, FeatureCollection, LineString

logger = getLogger(__name__)

def LoadFromFile(file):
    workbook = load_workbook(file, read_only = True)
    #load values from xls
    for sheet in workbook:
        try:          
            data = _ParseSheet(sheet)
        except KeyboardInterrupt:
            logger.warning("KeyboardInterrupt")
            raise KeyboardInterrupt
        except e:
            logger.error("Exception reading source data: %s", str(e))
            return

    return data


def _ParseSheet(sheet):

    normalize = {'No' : 'number',
            'FID' : 'FID',
            'SILNICE' : 'road',
            'KOD_TR_KOM' : 'KOD_TR_KOM',
            'sirka' : 'maxwidth',
            'vyska' : 'maxheight',
            'nosnost' : 'maxweight',
            'kategorie_vozidla' : 'vehicle_category',
            'Y' : 'latitude',
            'X' : 'longitude'}


    data = []
    rows = sheet.rows
    first_row = [cell.value for cell in next(rows)]
    for row in rows:#get_tqdm(rows, self.SetState, desc="loading " + sheet.title, total=sheet.max_row):
        record = {}
        for key, cell in zip(first_row, row):
            if cell.data_type == 's':
                record[normalize.get(key, key)] = cell.value.strip()
            else:
                record[normalize.get(key, key)] = cell.value

        data.append(record)
    return data
    

def ExportGeoJSON(data):
    features = []
    for point in data:
        features.append( Feature(id=point['number'], geometry=Point((float(point['longitude']), float(point['latitude'])))) )
    return FeatureCollection(features)